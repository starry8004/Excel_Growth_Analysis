import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
from datetime import datetime

class ExcelProcessor:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("셀링하니 상품 분석기")
        self.root.geometry("500x300")
        self.setup_gui()

    def setup_gui(self):
        # 메인 프레임
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # 상태 레이블
        self.status_label = ttk.Label(self.main_frame, text="파일을 선택해주세요")
        self.status_label.pack(pady=10)

        # 진행바
        self.progress_bar = ttk.Progressbar(self.main_frame, length=400, mode='determinate')
        self.progress_bar.pack(pady=10)

        # 버튼 프레임
        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.pack(pady=20)

        # 파일 선택 버튼
        self.select_button = ttk.Button(self.button_frame, text="파일 선택", command=self.select_file)
        self.select_button.pack(side=tk.LEFT, padx=5)

        # 분석 옵션 프레임
        self.options_frame = ttk.LabelFrame(self.main_frame, text="분석 옵션", padding="10")
        self.options_frame.pack(fill=tk.X, pady=10)

        # 분석 옵션 버튼
        self.growth_button = ttk.Button(self.options_frame, text="성장 상품 분석", 
                                      command=lambda: self.process_file("growth"))
        self.growth_button.pack(side=tk.LEFT, padx=5)
        
        self.rapid_growth_button = ttk.Button(self.options_frame, text="급성장 상품 분석", 
                                            command=lambda: self.process_file("rapid_growth"))
        self.rapid_growth_button.pack(side=tk.LEFT, padx=5)

        self.options_frame.pack_forget()  # 초기에는 숨김

    def select_file(self):
        self.filename = filedialog.askopenfilename(
            title="엑셀 파일을 선택하세요",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if self.filename:
            self.status_label.config(text="파일이 선택되었습니다.")
            self.options_frame.pack(fill=tk.X, pady=10)

    def get_column_indices(self, sheet):
        headers = {}
        for idx, cell in enumerate(sheet[1], 1):
            headers[cell.value] = idx
        return headers

    def convert_to_float(self, value):
        if isinstance(value, (int, float)):
            return float(value)
        elif isinstance(value, str):
            try:
                return float(value.replace(',', ''))
            except ValueError:
                return 0.0
        return 0.0

    def process_file(self, analysis_type):
        if not hasattr(self, 'filename'):
            messagebox.showerror("에러", "파일을 먼저 선택해주세요.")
            return

        try:
            self.status_label.config(text="파일 처리 중...")
            self.progress_bar['value'] = 20
            self.root.update()

            # 엑셀 파일 읽기
            wb = openpyxl.load_workbook(self.filename)
            sheet = wb.active
            
            # 컬럼 인덱스 찾기
            headers = self.get_column_indices(sheet)
            required_columns = ['성장성', '검색량', '쇼핑성키워드', '경쟁률', '키워드', 
                              '카테고리전체', '광고경쟁강도', '계절성']
            
            for col in required_columns:
                if col not in headers:
                    messagebox.showerror("에러", f"필수 컬럼 '{col}'을 찾을 수 없습니다.")
                    return

            self.progress_bar['value'] = 40
            self.root.update()

            # 새 워크북 생성
            new_wb = openpyxl.Workbook()
            new_sheet = new_wb.active
            
            # 헤더 작성
            new_headers = ['순서_검색량순', '키워드', '카테고리전체', '검색량', '경쟁률', 
                          '광고경쟁강도', '계절성']
            for col, header in enumerate(new_headers, 1):
                new_sheet.cell(1, col, header)

            # 데이터 필터링 및 정렬을 위한 임시 리스트
            filtered_data = []
            total_rows = sheet.max_row - 1

            for row in range(2, sheet.max_row + 1):
                self.progress_bar['value'] = 40 + (row/total_rows * 30)
                self.root.update()

                growth = self.convert_to_float(sheet.cell(row, headers['성장성']).value)
                search_volume = self.convert_to_float(sheet.cell(row, headers['검색량']).value)
                is_shopping = str(sheet.cell(row, headers['쇼핑성키워드']).value).lower() == 'true'
                competition = self.convert_to_float(sheet.cell(row, headers['경쟁률']).value)

                meets_criteria = False
                if analysis_type == "growth":
                    meets_criteria = (growth >= 0 and search_volume >= 8000 and 
                                    is_shopping and competition < 4)
                else:  # rapid_growth
                    meets_criteria = (growth >= 0.15 and search_volume >= 10000 and 
                                    is_shopping)

                if meets_criteria:
                    row_data = [
                        sheet.cell(row, headers['키워드']).value,
                        sheet.cell(row, headers['카테고리전체']).value,
                        search_volume,
                        competition,
                        sheet.cell(row, headers['광고경쟁강도']).value,
                        sheet.cell(row, headers['계절성']).value
                    ]
                    filtered_data.append(row_data)

            # 검색량 기준 정렬
            filtered_data.sort(key=lambda x: x[2], reverse=True)

            # 정렬된 데이터 쓰기
            for idx, row_data in enumerate(filtered_data, 1):
                self.progress_bar['value'] = 70 + (idx/len(filtered_data) * 20)
                self.root.update()
                
                new_sheet.cell(idx + 1, 1, idx)  # 순서_검색량순
                for col, value in enumerate(row_data, 2):
                    new_sheet.cell(idx + 1, col, value)

            # 파일 저장
            suffix = "성장" if analysis_type == "growth" else "급성장"
            current_time = datetime.now().strftime('%Y-%m-%d_%H-%M')
            file_dir = os.path.dirname(self.filename)
            file_name = os.path.splitext(os.path.basename(self.filename))[0]
            output_file = os.path.join(file_dir, f"{file_name}_{suffix}_{current_time}.xlsx")
            
            new_wb.save(output_file)

            self.progress_bar['value'] = 100
            self.root.update()

            messagebox.showinfo("완료", 
                f"처리가 완료되었습니다!\n\n"
                f"- 전체 데이터: {total_rows}행\n"
                f"- 필터링된 데이터: {len(filtered_data)}행\n\n"
                f"저장 경로:\n{output_file}")

        except Exception as e:
            messagebox.showerror("에러", f"처리 중 오류가 발생했습니다:\n{str(e)}")
        
        finally:
            self.progress_bar['value'] = 0
            self.status_label.config(text="파일을 선택해주세요")

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = ExcelProcessor()
    app.run()